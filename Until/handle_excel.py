#coding=utf-8

import openpyxl
import sys,os,json
from Until.handle_ini import handle_init

CasePath = os.path.abspath('../Case/')

class HandleExcel:
    def load_excle(self):
        '''
        加载excel
        :return:
        '''
        open_excel = openpyxl.load_workbook(CasePath + r'\clear_app.xlsx')
        return open_excel

    def get_sheet_data(self,index=None):
        '''
        加载所有sheet内容
        :param index:sheet下标
        :return:
        '''
        sheet_name = self.load_excle().sheetnames
        if index == None:
            index = 0
        data =self.load_excle()[sheet_name[index]]
        return data

    def get_rows(self):
        '''
        获取总行数
        :return:
        '''
        data = self.get_sheet_data().max_row
        return data

    def get_rows_number(self,case_id):
        '''
        通过case_id获取行号
        :param case_id:
        :return:
        '''
        num = 1
        cols_data =self.get_column_value()
        for col_data in cols_data:
            if case_id ==col_data:
                return num
            num += 1
        return num


    def get_rows_value(self,row):
        '''
        获取某一行的数据
        :param row: 第几行
        :return:
        '''
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def get_columns(self):
        '''
        获取总列数
        :return:
        '''
        columns = self.get_sheet_data().max_column
        return columns

    def get_column_value(self,key=None):
        '''
        获取某一列的数据
        :param cell: 第几列(A\B\C\D\...)
        :return:
        '''
        columns_list = []
        if key is None:
            key = 'A'
        columns_list_data =self.get_sheet_data()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_cell_value(self,row,cols):
        '''
        获取单元格的内容
        :param row:
        :param cols:
        :return:
        '''
        data =self.get_sheet_data().cell(row=row,column=cols).value
        return data

    def excel_wirte_data(self,row,cols,value):
        '''
        写入数据
        :param row:行
        :param cols: 列
        :param value: 数据
        :return:
        '''
        wb = self.load_excle()
        wr = wb.active
        wr.cell(row,cols,value)
        wb.save(CasePath + r'\clear_app.xlsx')

excel_data = HandleExcel()

if __name__ == '__main__':
    hdx = HandleExcel()
    #打印加载所有sheet内容
    # print(hdx.get_sheet_data())         #----------<Worksheet "clear_app">
    #获取第一行的数据
    # print(hdx.get_rows_value(1))       #---------['case编号', '是否执行', 'url', 'method', 'data', 'cookie操作', '预期结果', 'result']
    #获取总行数
    # print(hdx.get_rows())              #-----------------4
    #获取总列数
    # print(hdx.get_columns())              #---------8
    #获取第一列的数据
    # print(hdx.get_column_value('J'))       #-----['case编号', 'case_001', None, None]
    # 获取单元格的数据
    print(hdx.get_cell_value(1,2))           #case编号
    #向单元格写入数据
    # print(hdx.excel_wirte_data(3,1,'case_002'))     #写入成功，但是是不会反悔
    # print(type(hdx.get_cell_value(2,10)))
    # b = json.loads(hdx.get_cell_value(2,10))
    # print(b)
    # print(type(b))
    # print(hdx.get_rows_number('imooc_003'))


